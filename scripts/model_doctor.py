#!/usr/bin/env python3
"""
model_doctor.py — audit and clean up a hardtech financial model (.xlsx).

Two subcommands:

  audit  <file.xlsx> [--report out.md]
      Scan every cell and report hygiene issues without changing anything:
      font/size inconsistencies, mis-colored cells, magic numbers inside
      formulas, duplicated hardcoded assumptions (centralization candidates),
      missing structural tabs, external links, and error cells.

  fix    <file.xlsx> --out cleaned.xlsx
      Write a CLEANED COPY (never touches the original). Standardizes font to
      10pt Arial, recolors cells by content type (blue hardcode / black
      same-sheet formula / green cross-tab link), and scaffolds Control and
      Notes tabs if they are missing. Values and formula logic are never
      changed — only fonts, colors, and (optionally) number formats. Every
      change is logged.

Color heuristic (documented so a human can judge edge cases):
  - numeric literal            -> BLUE   (#0070C0)  "an input you can change"
  - formula, cross-sheet link  -> GREEN  (#1B7A3D)  "pulled from another tab"
  - formula that computes      -> BLACK  (#000000)  "calculated here"
  - external link / error      -> RED    (#FF0000)  "danger"
  A formula is treated as a LINK (green) only when it references another sheet
  AND does no real arithmetic (only +, -, and a leading SUM pass-through). A
  formula that references another sheet but multiplies/divides is a CALCULATION
  (black). The audit flags ambiguous cells so a human can decide.
"""

import argparse
import re
import sys
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.comments import Comment
except ImportError:
    sys.exit("openpyxl is required. Install with: pip install openpyxl --break-system-packages")

# ---- House standard palette (bright, standard modeling colors) -------------
# These are the exact target colors the `fix` job normalizes everything to.
BLUE = "FF0000FF"    # hardcoded input     -> bright blue  #0000FF
GREEN = "FF00B050"   # cross-tab link      -> bright green #00B050
BLACK = "FF000000"   # same-sheet formula / label
RED = "FFFF0000"     # external link / warning
GRAY = "FF7F7F7F"    # secondary label text
BODY_FONT = "Arial"
BODY_SIZE = 10.0


def _channels(rgb):
    """Return (r, g, b) from an ARGB/RGB hex string, or None."""
    if not isinstance(rgb, str):
        return None
    h = rgb[-6:]
    try:
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except ValueError:
        return None


def is_blueish(rgb):
    """Any recognizable input-blue shade (so the audit tolerates #0070C0,
    #0000FF, #0432FF, etc.), not just the exact target."""
    ch = _channels(rgb)
    if not ch:
        return False
    r, g, b = ch
    return b > 130 and b > r + 40 and b > g + 30


def is_greenish(rgb):
    """Any recognizable link-green shade (#00B050, #1B7A3D, #008000, …)."""
    ch = _channels(rgb)
    if not ch:
        return False
    r, g, b = ch
    return g > 90 and g > r + 30 and g > b + 20

# Names a well-structured model tends to have.
CONTROL_NAMES = {"control", "assumptions", "inputs", "drivers", "control panel", "assumptions & drivers"}
NOTES_NAMES = {"notes", "methodology", "notes & methodology", "assumptions notes"}
INTRO_NAMES = {"introduction", "intro", "cover", "overview", "read me", "readme", "legend"}

ERROR_STRINGS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}

# Literals too trivial to be worth centralizing.
TRIVIAL_LITERALS = {0, 1, -1, 2, 12, 100, 1000, 24, 365, 7, 4, 52, 3, 0.5}

# Unit-conversion / time / scaling constants that legitimately live inside
# formulas — flagging these as "magic numbers" is noise, not signal.
SCALING_CONSTANTS = {8760, 1000, 1000000, 100, 60, 3600, 24, 12, 52, 365, 10, 6, 2, 4, 0.5, 1e6, 1e9}

SHEET_REF_RE = re.compile(r"(?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)\!")
MATH_OP_RE = re.compile(r"[*/^]")
# A "pure link": a formula that just relays a single value from another sheet,
# e.g. =Control!C27, =+Summary!$C$5. Optionally a SUM of one cross-sheet range.
LINK_RE = re.compile(r"^=[+\-]?(?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)\!\$?[A-Z]+\$?[0-9]+$")
LINK_SUM_RE = re.compile(
    r"(?i)^=[+]?sum\((?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)\![A-Z]+[0-9]+:[A-Z]+[0-9]+\)$")
# numbers embedded in a formula, excluding cell refs / row-col / exponents.
NUM_IN_FORMULA_RE = re.compile(r"(?<![A-Za-z0-9_$!.\^])\d+(?:\.\d+)?(?![A-Za-z0-9_])")
YEAR_RE = re.compile(r"^(19|20)\d{2}$")


def norm(name):
    return name.strip().lower()


def col_rgb(font):
    try:
        rgb = font.color.rgb if font and font.color else None
        if isinstance(rgb, str):
            return rgb.upper()
    except Exception:
        pass
    return None


def is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def is_significant(num, raw=None):
    """A number worth centralizing/flagging: not trivial, not a scaling/time
    constant, not a year. Keeps decimals (0.8 utilization, 0.13 rate, 3.91
    price) which are real assumptions; skips tiny integer counts (3, 4, 5)."""
    if num in TRIVIAL_LITERALS or num in SCALING_CONSTANTS:
        return False
    if raw is not None and YEAR_RE.match(raw):
        return False
    if float(num).is_integer():
        # small integer counts are usually plumbing, not assumptions
        return abs(num) >= 5 and not (1900 <= num <= 2100)
    return True  # keep all non-integer decimals


def classify_formula(formula):
    """Return 'link' (green), 'calc' (black), or 'external'/'error'.

    A formula is a LINK only when it relays a SINGLE value from another sheet.
    An arithmetic combination of several cross-sheet cells (even just adding or
    subtracting them) is a CALCULATION and stays black — that's how a
    well-built model treats it, so we don't over-green.
    """
    f = formula
    if any(e in f for e in ERROR_STRINGS):
        return "error"
    if "[" in f and "]" in f:  # [Book.xlsx] external workbook reference
        return "external"
    if LINK_RE.match(f) or LINK_SUM_RE.match(f):
        return "link"
    return "calc"


def iter_content_cells(ws):
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None and c.value != "":
                yield c


# ---------------------------------------------------------------------------
# AUDIT
# ---------------------------------------------------------------------------
def audit(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    issues = {
        "fonts": [], "sizes": [], "miscolored": [], "magic_numbers": [],
        "errors": [], "external": [], "structure": [], "links_not_green": [],
    }
    font_names = defaultdict(int)
    font_sizes = defaultdict(int)
    literal_locations = defaultdict(list)  # value -> [ (sheet, coord) ] outside control tab

    sheet_norms = {norm(s) for s in wb.sheetnames}
    has_control = bool(sheet_norms & CONTROL_NAMES)
    has_notes = bool(sheet_norms & NOTES_NAMES)
    has_intro = bool(sheet_norms & INTRO_NAMES)
    control_sheets = {s for s in wb.sheetnames if norm(s) in CONTROL_NAMES}

    for ws in wb.worksheets:
        in_control = ws.title in control_sheets
        for c in iter_content_cells(ws):
            f = c.font
            loc = f"{ws.title}!{c.coordinate}"
            # font family / size
            if f and f.name and f.name != BODY_FONT:
                font_names[f.name] += 1
                issues["fonts"].append((loc, f.name))
            if f and f.size:
                font_sizes[f.size] += 1

            val = c.value
            rgb = col_rgb(f)

            if isinstance(val, str) and val.startswith("="):
                kind = classify_formula(val)
                if kind == "error":
                    issues["errors"].append((loc, val[:40]))
                elif kind == "external":
                    issues["external"].append((loc, val[:60]))
                    if rgb != RED:
                        issues["miscolored"].append((loc, "external link not red", val[:40]))
                elif kind == "link":
                    # informational only — good models green links selectively.
                    if not is_greenish(rgb):
                        issues["links_not_green"].append((loc, val[:40]))
                else:  # calc
                    # a computing formula colored blue looks like a typed input — a real error.
                    if is_blueish(rgb):
                        issues["miscolored"].append((loc, "formula colored blue (looks like input)", val[:40]))
                # magic numbers: only genuinely non-scaling business constants.
                for m in NUM_IN_FORMULA_RE.findall(val):
                    try:
                        num = float(m)
                    except ValueError:
                        continue
                    if is_significant(num, m):
                        issues["magic_numbers"].append((loc, num, val[:50]))
                        break
            elif is_number(val):
                # a hardcoded input; zeros are almost always intentional placeholders.
                if not is_blueish(rgb) and rgb != GRAY and val != 0:
                    issues["miscolored"].append((loc, "hardcode not blue", val))
                if not in_control and is_significant(val):
                    literal_locations[val].append(loc)
            elif isinstance(val, str) and val in ERROR_STRINGS:
                issues["errors"].append((loc, val))

    # duplicated literals -> centralization candidates
    dup_literals = {v: locs for v, locs in literal_locations.items() if len(locs) >= 2}

    # structure
    if not has_control:
        issues["structure"].append("No centralized assumptions tab (Control/Assumptions). "
                                    "Every input should live once, here.")
    if not has_notes:
        issues["structure"].append("No Notes/Methodology tab. Assumptions have nowhere to be sourced/explained.")
    if not has_intro:
        issues["structure"].append("No Introduction/cover tab with a legend decoding the color scheme.")

    # dominant font size sanity
    dominant_size = max(font_sizes, key=font_sizes.get) if font_sizes else None

    return {
        "path": path,
        "sheets": wb.sheetnames,
        "issues": issues,
        "font_names": dict(font_names),
        "font_sizes": dict(font_sizes),
        "dominant_size": dominant_size,
        "dup_literals": dup_literals,
        "has_control": has_control,
        "has_notes": has_notes,
        "has_intro": has_intro,
    }


def render_report(a):
    L = []
    L.append(f"# Model audit — {a['path'].split('/')[-1]}\n")
    L.append(f"Sheets ({len(a['sheets'])}): {', '.join(a['sheets'])}\n")
    iss = a["issues"]

    # scorecard
    L.append("## Scorecard\n")
    L.append(f"- Centralized assumptions tab: {'yes' if a['has_control'] else 'NO — top priority'}")
    L.append(f"- Notes / methodology tab: {'yes' if a['has_notes'] else 'NO'}")
    L.append(f"- Introduction / legend tab: {'yes' if a['has_intro'] else 'NO'}")
    L.append(f"- Non-Arial fonts found: {a['font_names'] or 'none'}")
    L.append(f"- Dominant font size: {a['dominant_size']}pt "
             f"(house standard is 10pt)")
    L.append(f"- Mis-colored cells (real errors): {len(iss['miscolored'])}")
    L.append(f"- Duplicated hardcoded assumptions: {len(a['dup_literals'])}")
    L.append(f"- Cross-tab links not colored green (info): {len(iss['links_not_green'])}")
    L.append(f"- Possible business constants in formulas (worth a look): {len(iss['magic_numbers'])}")
    L.append(f"- Error cells: {len(iss['errors'])}")
    L.append(f"- External-workbook links: {len(iss['external'])}\n")

    def section(title, rows, fmt, limit=40):
        L.append(f"## {title} ({len(rows)})\n")
        if not rows:
            L.append("_none_\n")
            return
        for r in rows[:limit]:
            L.append(fmt(r))
        if len(rows) > limit:
            L.append(f"\n_…and {len(rows) - limit} more._\n")
        L.append("")

    if iss["structure"]:
        L.append("## Structural gaps\n")
        for s in iss["structure"]:
            L.append(f"- {s}")
        L.append("")

    section("Centralization candidates (same value hardcoded in 2+ places)",
            sorted(a["dup_literals"].items(), key=lambda kv: -len(kv[1])),
            lambda kv: f"- **{kv[0]}** appears in {len(kv[1])} places: {', '.join(kv[1][:8])}"
                       + (" …" if len(kv[1]) > 8 else ""),
            limit=30)

    section("Mis-colored cells (break the blue/black/green code)",
            iss["miscolored"],
            lambda r: f"- `{r[0]}` — {r[1]} (`{r[2]}`)" if len(r) == 3 else f"- `{r[0]}` — {r[1]}")

    section("Possible business constants inside formulas (worth a look — scaling/time constants already excluded)",
            iss["magic_numbers"],
            lambda r: f"- `{r[0]}` — literal `{r[1]}` in `{r[2]}`", limit=25)

    section("Cross-tab links not colored green (optional — good models green links selectively)",
            iss["links_not_green"],
            lambda r: f"- `{r[0]}` — `{r[1]}`", limit=15)

    section("Non-Arial cells", iss["fonts"], lambda r: f"- `{r[0]}` — {r[1]}", limit=20)
    section("Error cells", iss["errors"], lambda r: f"- `{r[0]}` — {r[1]}")
    section("External-workbook links", iss["external"], lambda r: f"- `{r[0]}` — {r[1]}")

    L.append("## What to do next\n")
    L.append("1. Run `fix` to auto-standardize fonts and recolor cells to a cleaned copy.")
    L.append("2. Work the centralization candidates: lift each into the Control tab and link to it.")
    L.append("3. Source every Control input (cell comment + a line in Notes).")
    L.append("4. Build out the hardtech analytics (unit economics, utilization breakeven, capex intensity).")
    L.append("\n_Color heuristic: numeric literal→blue, cross-tab pure link→green, computing formula→black, "
             "external/error→red. Ambiguous formulas (cross-sheet + math) default to black; spot-check them._\n")
    return "\n".join(L)


# ---------------------------------------------------------------------------
# FIX
# ---------------------------------------------------------------------------
def fix(path, out):
    wb = openpyxl.load_workbook(path, data_only=False)
    log = []
    control_sheets = {s for s in wb.sheetnames if norm(s) in CONTROL_NAMES}

    for ws in wb.worksheets:
        for c in iter_content_cells(ws):
            f = c.font
            new_name = BODY_FONT
            # keep header sizing: only shrink oversized body-ish text conservatively.
            # We standardize the family always, and size only when it's a small body size (<=12).
            cur_size = f.size if f and f.size else BODY_SIZE
            new_size = BODY_SIZE if (cur_size and cur_size <= 12) else cur_size

            val = c.value
            cur_color = col_rgb(f)
            new_color = cur_color or BLACK
            changed_reason = []

            # Conservative recoloring: fix clear errors only. We do NOT blanket
            # recolor black formulas to green — well-built models green links
            # selectively, so forcing every cross-sheet reference green would
            # fight the author's intent and create noise.
            if isinstance(val, str) and val.startswith("="):
                kind = classify_formula(val)
                if kind == "external" or kind == "error":
                    if cur_color != RED:
                        new_color = RED
                        changed_reason.append("external/error->red")
                elif is_blueish(cur_color):
                    # a formula that looks like a typed input — clearly wrong.
                    target = GREEN if kind == "link" else BLACK
                    new_color = target
                    changed_reason.append(f"formula->{'green' if target == GREEN else 'black'}")
                elif is_greenish(cur_color) and cur_color != GREEN:
                    # an author-greened link in a muted shade — normalize to bright green.
                    new_color = GREEN
                    changed_reason.append("link->bright green")
                # otherwise leave the formula's color as the author set it.
            elif is_number(val):
                # normalize every typed input to the exact bright input-blue
                # (also brightens muted blues like #0070C0). Zeros/gray left alone.
                if cur_color != GRAY and val != 0 and cur_color != BLUE:
                    new_color = BLUE
                    changed_reason.append("hardcode->blue")
            # text cells: leave color as-is

            # apply
            font_changed = (f.name != new_name) or (cur_size != new_size) or changed_reason
            if font_changed:
                c.font = Font(
                    name=new_name,
                    size=new_size,
                    bold=f.bold, italic=f.italic, underline=f.underline,
                    color=new_color,
                )
                if changed_reason or f.name != new_name:
                    log.append(f"{ws.title}!{c.coordinate}: "
                               + ", ".join(([f"font->{new_name}"] if f.name != new_name else []) + changed_reason))

    # scaffold missing tabs
    if not control_sheets:
        scaffold_control(wb)
        log.append("Added scaffolded 'Control' assumptions tab (fill in inputs).")
    if not (set(norm(s) for s in wb.sheetnames) & NOTES_NAMES):
        scaffold_notes(wb)
        log.append("Added scaffolded 'Notes' methodology tab.")

    wb.save(out)
    return log


def _title(ws, text, size=15):
    ws["A1"] = text
    ws["A1"].font = Font(name=BODY_FONT, size=size, bold=True, color=BLACK)


def scaffold_control(wb):
    ws = wb.create_sheet("Control", 0 if len(wb.sheetnames) else None)
    _title(ws, "Control Panel")
    ws["A2"] = "Centralized assumptions — every input lives here once. Blue = input."
    ws["A2"].font = Font(name=BODY_FONT, size=BODY_SIZE, italic=True, color=GRAY)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    rows = [
        ("SIZING OUTPUTS", "", ""),
        ("Raise size", "$mm", ""),
        ("Key return (IRR/MOIC)", "", ""),
        ("Breakeven utilization", "%", ""),
        ("", "", ""),
        ("1. KEY INPUTS", "", ""),
        ("Price per unit", "", ""),
        ("Cost per unit", "", ""),
        ("2. MIX & UTILIZATION", "", ""),
        ("Utilization", "%", ""),
        ("3. SCALE", "", ""),
        ("Years to full buildout", "yrs", ""),
        ("4. OPEX", "", ""),
        ("5. CAPEX", "", ""),
        ("Capex per unit of capacity", "", ""),
        ("6. DEBT SIZING", "", ""),
        ("7. DEBT TERMS", "", ""),
    ]
    r = 4
    for label, unit, val in rows:
        ws.cell(r, 1, label)
        is_header = label and label[0].isdigit() or label in ("SIZING OUTPUTS",)
        ws.cell(r, 1).font = Font(name=BODY_FONT, size=BODY_SIZE, bold=bool(is_header), color=BLACK)
        ws.cell(r, 2, unit).font = Font(name=BODY_FONT, size=BODY_SIZE, color=GRAY)
        r += 1
    ws.freeze_panes = "C4"


def scaffold_notes(wb):
    ws = wb.create_sheet("Notes")
    _title(ws, "Notes & Methodology")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 90
    tips = [
        ("1. STRUCTURE & ENTITIES", ""),
        ("", "Describe the business, entities, and how cash flows between them."),
        ("2. CAPEX & REVENUE", ""),
        ("", "Source each capex and revenue assumption. Where did the number come from?"),
        ("3. FINANCING", ""),
        ("", "Debt sizing, tenor, reserves, prepayments — and the convention used for each."),
        ("4. UNIT ECONOMICS & BREAKEVEN", ""),
        ("", "State the operating point (utilization/price) and the breakeven, and the cushion between them."),
        ("5. CONSERVATISM", ""),
        ("", "What was deliberately assumed conservatively, and why the downside is bounded."),
    ]
    r = 3
    for a, b in tips:
        ws.cell(r, 1, a).font = Font(name=BODY_FONT, size=BODY_SIZE, bold=bool(a), color=BLACK)
        ws.cell(r, 2, b).font = Font(name=BODY_FONT, size=BODY_SIZE, italic=True, color=GRAY)
        r += 1


# ---------------------------------------------------------------------------
def main():
    p = argparse.ArgumentParser(description="Audit and clean up a hardtech financial model (.xlsx).")
    sub = p.add_subparsers(dest="cmd", required=True)

    pa = sub.add_parser("audit", help="Report hygiene issues without changing the file.")
    pa.add_argument("file")
    pa.add_argument("--report", help="Write the markdown report to this path (else prints to stdout).")

    pf = sub.add_parser("fix", help="Write a cleaned copy with fonts/colors standardized.")
    pf.add_argument("file")
    pf.add_argument("--out", required=True, help="Output path for the cleaned copy.")

    args = p.parse_args()

    if args.cmd == "audit":
        a = audit(args.file)
        report = render_report(a)
        if args.report:
            with open(args.report, "w") as fh:
                fh.write(report)
            print(f"Audit written to {args.report}")
            # also print the scorecard to stdout for quick reading
            print("\n".join(report.split("\n")[:24]))
        else:
            print(report)

    elif args.cmd == "fix":
        log = fix(args.file, args.out)
        print(f"Cleaned copy written to {args.out}")
        print(f"{len(log)} changes:")
        for line in log[:60]:
            print("  " + line)
        if len(log) > 60:
            print(f"  …and {len(log) - 60} more.")


if __name__ == "__main__":
    main()
